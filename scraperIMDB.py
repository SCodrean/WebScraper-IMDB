#import libraries
from bs4 import BeautifulSoup       #webscraping and parsing HTML and XML documents. Extracts data from web pages
import requests                     #Used for making HTTP requests to communicate with webservers. Allows you to interact with webpages, API's, etc.
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top 250 Rated Movies'
print(excel.sheetnames)
sheet.append(['Film Rank', 'Film Name', 'Release Date', 'IMDB RATING'])


try:
    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()   #raise_for_status() throws error if URL not found

    #method allws user to navigate and extrade data from HTML structure of web page.
    soup = BeautifulSoup(source.text,'html.parser')     #source.text = HMTL content of web page obtained from requests.get() method. Contains raw data of page
                                                        #html.parser, parses HTML content used by BeautifulSoup.   

    #.find() finds the first match, with tag 'tbody' and has class "lister-list". find_all() finds all the entries with tr tag.
    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    for movie in movies:
        name = movie.find('td', class_ = "titleColumn").a.text
        rank = movie.find('td', class_ = "titleColumn").get_text(strip = True).split('.')[0]
        year = movie.find('td', class_ = "titleColumn").span.text.strip('()')
        rating = movie.find('td', class_ = "ratingColumn imdbRating").strong.text
        print(rank, name, year, rating) 
        sheet.append([rank, name, year, rating])
        
        

except Exception as e:
    print(e)

excel.save('IMDB MOVIE STATS.xlsx')
   